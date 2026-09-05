#!/usr/bin/env python3
"""
Merge scraper output into PSK50_program_data.xlsx.

    python psk50_merge.py

Reads psk50_presentations.csv (stage 1) and, if present, psk50_abstracts.csv
(stage 2), joins them by pid, and writes three sheets:

  Talks     one row per podium talk
  Chairs    who chairs which room, when — derived from the talk rows
  Rooms     which track occupies which room on which day
"""
import argparse, csv, os, sys, collections
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR = Font(name=ARIAL, bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="2F5597")
BODY = Font(name=ARIAL)

TALK_COLS = ["session", "type", "date", "start", "end", "room", "chair",
             "presenter", "affiliation", "title", "abstract_no", "coauthors",
             "abstract", "pid", "url"]
TALK_W = [9, 10, 11, 7, 7, 12, 20, 24, 30, 60, 13, 34, 80, 7, 56]


def load(path):
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def sheet(wb, name, header, rows, widths):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(header)
    for r in rows:
        ws.append([r.get(c, "") for c in header] if isinstance(r, dict) else r)
    for c in range(1, len(header) + 1):
        ws.cell(1, c).font, ws.cell(1, c).fill = HDR, FILL
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def sort_key(t):
    return (t.get("date") or "9999", t.get("start") or "99:99",
            t.get("room") or "", t.get("session") or "")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default="PSK50_program_data.xlsx")
    ap.add_argument("--presentations", default="psk50_presentations.csv")
    ap.add_argument("--abstracts", default="psk50_abstracts.csv")
    args = ap.parse_args()

    talks = load(args.presentations)
    if not talks:
        sys.exit(f"{args.presentations} not found or empty — run psk50_scrape.py first")

    abstracts = {a["pid"]: a for a in load(args.abstracts)}
    for t in talks:
        a = abstracts.get(t["pid"], {})
        for k in ("abstract_no", "coauthors", "abstract"):
            t[k] = a.get(k, "")
    talks.sort(key=sort_key)

    # chairs: one row per contiguous chair stint in a room on a day
    stints = collections.defaultdict(list)
    for t in talks:
        if t.get("chair") and t.get("date"):
            stints[(t["date"], t["room"], t["session"], t["chair"])].append(t)
    chairs = sorted(
        [[d, room, sess, chair, min(x["start"] for x in ts),
          max(x["end"] for x in ts), len(ts)]
         for (d, room, sess, chair), ts in stints.items()],
        key=lambda r: (r[0], r[4], r[1]))

    # rooms: which track sat in which room, per day
    occ = collections.defaultdict(set)
    for t in talks:
        if t.get("date") and t.get("room"):
            occ[(t["date"], t["room"])].add(t["session"])
    rooms = sorted([[d, room, ", ".join(sorted(s))] for (d, room), s in occ.items()])

    # Create the workbook on first run rather than demanding one already exists.
    if os.path.exists(args.workbook):
        wb = load_workbook(args.workbook)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        os.makedirs(os.path.dirname(args.workbook) or ".", exist_ok=True)
    sheet(wb, "Talks", TALK_COLS, talks, TALK_W)
    sheet(wb, "Chairs", ["date", "room", "track", "chair", "from", "to", "talks"],
          chairs, [11, 12, 9, 24, 8, 8, 8])
    sheet(wb, "Rooms", ["date", "room", "tracks"], rooms, [11, 12, 22])
    wb.save(args.workbook)

    types = collections.Counter(t.get("type") for t in talks)
    print(f"{len(talks)} talks -> '{args.workbook}'")
    print("  by type    :", ", ".join(f"{k} {v}" for k, v in types.most_common()))
    print(f"  with chair : {sum(1 for t in talks if t.get('chair'))}")
    print(f"  abstracts  : {sum(1 for t in talks if t.get('abstract'))}")
    print(f"  chair stints: {len(chairs)} · rooms: {len({r[1] for r in rooms})}")


if __name__ == "__main__":
    main()
